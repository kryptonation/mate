### app/ledger/utils.py

# Standard library import
import os
from datetime import datetime, date, timedelta, time, timezone
from collections import defaultdict
from pathlib import Path
from io import BytesIO

# Third party imports
from jinja2 import Environment, FileSystemLoader
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Local imports
from app.utils.logger import get_logger
from app.ledger.models import LedgerEntry
from app.ledger.schemas import LedgerSourceType
from app.utils.s3_utils import s3_utils
from app.utils.exporter.pdf_exporter import PDFExporter

logger = get_logger(__name__)

def get_pay_window(pay_day: str, reference_date: date) -> tuple[datetime, datetime]:
    """Get pay window for a given pay day and reference date"""
    try:
        weekday_map = {
            "Monday": 0, "Tuesday": 1, "Wednesday": 2,
            "Thursday": 3, "Friday": 4, "Saturday": 5,
            "Sunday": 6
        }

        target_dow = weekday_map[pay_day]
        ref_dow = reference_date.weekday()

        days_delta = (ref_dow - target_dow) % 7
        period_start = datetime.combine(reference_date - timedelta(days=days_delta), time(5, 0))
        period_end = period_start + timedelta(days=7) - timedelta(minutes=1)

        return period_start, period_end
    except Exception as e:
        logger.error("Error getting pay window: %s", e, exc_info=True)
        raise e
    
def summarize_ledger_entries(entries: list[LedgerEntry]) -> dict:
    """Summarize ledger entries"""
    try:
        summary = defaultdict(float)

        for e in entries:
            if e.debit:
                if "lease" in e.description.lower():
                    summary["lease_due"]  += float(e.amount)
                elif e.source_type == LedgerSourceType.EZPASS:
                    summary["ezpass_due"] += float(e.amount)
                elif e.source_type == LedgerSourceType.PVB:
                    summary["pvb_due"] += float(e.amount)
                elif e.source_type == LedgerSourceType.MANUAL_FEE:
                    summary["manual_fee"] += float(e.amount)
                elif "mta" in e.description.lower():
                    summary["mta"] += float(e.amount)
                elif "tif" in e.description.lower():
                    summary["tif"] += float(e.amount)
                elif "cps" in e.description.lower():
                    summary["cps"] += float(e.amount)
                elif "aaf" in e.description.lower():
                    summary["aaf"] += float(e.amount)
            else:
                if "trip fare (card)" in e.description.lower():
                    summary["cc_earnings"] += float(e.amount)
                elif "trip fare (cash)" in e.description.lower():
                    summary["cash_earnings"] += float(e.amount)
                elif "trip" in e.description.lower():
                    summary["trips"] += float(e.amount)
                elif "cash payment" in e.description.lower():
                    summary["cash_paid"] += float(e.amount)

            summary["total_dues"] = {
                summary["lease_due"] + summary["ezpass_due"] + summary["pvb_due"] +
                summary["manual_fee"] + summary["mta"] + summary["tif"] + summary["cps"] +
                summary["aaf"]
            }

            return summary
    except Exception as e:
        logger.error("Error summarizing ledger entries: %s", e, exc_info=True)
        raise e


def _safe_get_value(obj, attr_name, default=""):
    """Safely extract value from object, handling SQLAlchemy objects"""
    try:
        if hasattr(obj, attr_name):
            value = getattr(obj, attr_name)
            # Handle SQLAlchemy state objects
            if hasattr(value, '_sa_instance_state'):
                return str(value) if value is not None else default
            # Handle SQLAlchemy Column objects
            if hasattr(value, 'key'):
                return str(value) if value is not None else default
            return str(value) if value is not None else default
        return default
    except Exception:
        return default

def generate_dtr_html_doc(data: dict):
    """Generate DTR HTML document and upload to S3 using the new template."""
    try:
        TEMPLATE_PATH = Path(__file__).parent / "templates"
        env = Environment(loader=FileSystemLoader(TEMPLATE_PATH))
        # Use the new, more detailed template
        template = env.get_template("dtr_statement.html")

        # The 'data' dictionary is now the rich summary from generate_dtr_summary
        html_content = template.render(
            data=data, # Pass the entire data dictionary
            generated_at=datetime.now(timezone.utc)
        )

        html_bytes = BytesIO()
        html_bytes.write(html_content.encode("utf-8"))
        html_bytes.seek(0)

        key = f'dtr_receipts/driver_{data["driver"].id}/receipt_{data["receipt"].receipt_number}_{datetime.now(timezone.utc).timestamp()}.html'
        s3_utils.upload_file(html_bytes, key, content_type="text/html")
        return key
    except Exception as e:
        logger.error("Error generating DTR HTML document: %s", e, exc_info=True)
        raise e

def generate_dtr_pdf_doc(data: dict):
    """Generate DTR PDF document from the new HTML template and upload to S3."""
    try:
        TEMPLATE_PATH = Path(__file__).parent / "templates"
        env = Environment(loader=FileSystemLoader(TEMPLATE_PATH))
        template = env.get_template("dtr_statement.html")

        html_content = template.render(
            data=data,
            generated_at=datetime.now(timezone.utc)
        )

        pdf_bytes = BytesIO()
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_bytes, encoding='utf-8')
        
        if pisa_status.err:
            raise Exception(f"Error creating PDF: {pisa_status.err}")
        
        pdf_bytes.seek(0)
        key = f"dtr_receipts/driver_{data['driver'].id}/receipt_{data['receipt'].receipt_number}_{datetime.now(timezone.utc).timestamp()}.pdf"
        s3_utils.upload_file(pdf_bytes, key, content_type="application/pdf")
        return key
    except Exception as e:
        logger.error("Error generating DTR PDF document: %s", e, exc_info=True)
        raise e
    
def generate_dtr_excel_doc_styled(data: dict):
    """Generate styled DTR Excel document and upload to S3"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Driver Transaction Receipt"
        
        # Define styles
        header_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=14, color="FFFFFF")
        section_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        table_header_font = Font(bold=True)
        table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)
        
        # Set column widths
        column_widths = [20, 15, 15, 12, 12, 12, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Main header
        ws['A1'] = "Big Apple Taxi - Driver Transaction Receipt"
        ws.merge_cells('A1:I1')
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_alignment
        ws['A1'].border = border
        
        # Driver information with styling
        row = 3
        info_data = [
            ("Driver Name:", f"{data['driver'].first_name} {data['driver'].last_name}", 
             "Hack License:", data['driver'].tlc_license.tlc_license_number if hasattr(data['driver'], 'tlc_license') and data['driver'].tlc_license else "N/A"),
            ("Medallion:", data['medallion'].medallion_number if data['medallion'] else "N/A",
             "Cab:", data['vehicle'].registrations[0].plate_number if data['vehicle'] and data['vehicle'].registrations else "N/A"),
            ("Date Range:", f"{data['receipt'].period_start} to {data['receipt'].period_end}",
             "Generated:", str(data['receipt'].created_on))
        ]
        
        for info_row in info_data:
            ws[f'A{row}'] = info_row[0]
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = info_row[1]
            ws[f'D{row}'] = info_row[2]
            ws[f'D{row}'].font = bold_font
            ws[f'E{row}'] = info_row[3]
            row += 1
        
        # Trip Details Section
        row += 2
        ws[f'A{row}'] = "Trip Details"
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        
        row += 1
        # Trip headers
        headers = ['Trip Start Date', 'Trip End Date', 'Trip ID', 'Fare', 'Tip', 'TIF', 'MTA', 'CPS', 'AAF']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        logger.info(f"Generating Excel with {data}")
        # Trip data
        for trip in data['curb_trips']:
            row += 1
            trip_data = [
                f"{trip.start_date} {trip.start_time}",
                f"{trip.end_date} {trip.end_time}",
                f"{trip.record_id} {trip.period}",
                float(trip.trip_amount) if trip.trip_amount else 0,
                float(trip.tips) if trip.tips else 0,
                float(trip.distance_service) if trip.distance_service else 0,
                float(trip.ehail_fee) if trip.ehail_fee else 0,
                float(trip.congestion_fee) if trip.congestion_fee else 0,
                float(trip.airport_fee) if trip.airport_fee else 0
            ]
            
            for col, value in enumerate(trip_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = '$#,##0.00'
        
        # Ledger Entries Section
        row += 3
        ws[f'A{row}'] = "Ledger Entries"
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        
        row += 1
        # Ledger headers
        ledger_headers = ['ID', 'Date', 'Category', 'Amount', 'Type', 'Reference']
        for col, header in enumerate(ledger_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Ledger data
        for entry in data['ledgers']:
            row += 1
            ledger_data = [
                entry.id,
                str(entry.created_on),
                entry.source_type.value if hasattr(entry.source_type, 'value') else str(entry.source_type),
                float(entry.amount),
                "Debit" if entry.debit else "Credit",
                entry.source_id
            ]
            
            for col, value in enumerate(ledger_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 4 and isinstance(value, float):  # Amount column
                    cell.number_format = '$#,##0.00'
                    # Color code debit/credit
                    if ledger_data[4] == "Debit":
                        cell.font = Font(color="FF0000")  # Red for debit
                    else:
                        cell.font = Font(color="008000")  # Green for credit
        
        # Calculate totals
        total_trips = len(data['trips'])
        total_transactions = len(data['ledgers'])
        total_trip_amount = sum(float(trip.trip_amount) if trip.trip_amount else 0 for trip in data['trips'])
        total_tips = sum(float(trip.tips) if trip.tips else 0 for trip in data['trips'])
        total_credits = sum(float(entry.amount) for entry in data['ledgers'] if not entry.debit)
        total_debits = sum(float(entry.amount) for entry in data['ledgers'] if entry.debit)
        
        # Summary Section
        row += 3
        ws[f'A{row}'] = "Summary"
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        
        row += 1
        # Summary data
        summary_data = [
            ("Total Transactions:", total_transactions),
            ("Total Trips:", total_trips),
            ("Total Trip Amount:", total_trip_amount),
            ("Total Tips:", total_tips),
            ("Total Credits:", total_credits),
            ("Total Debits:", total_debits),
            ("Net Amount:", total_credits - total_debits),
            ("Returned to Driver:", float(data['receipt'].balance) if data['receipt'].balance else 0)
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = value
            if isinstance(value, float):
                ws[f'B{row}'].number_format = '$#,##0.00'
                # Highlight net amount and returned amount
                if "Net Amount" in label or "Returned to Driver" in label:
                    ws[f'A{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    ws[f'B{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Trip Summary by Date (if needed)
        if data['trips']:
            row += 2
            ws[f'A{row}'] = "Trip Summary by Date"
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].border = border
            
            row += 1
            # Date summary headers
            date_headers = ['Date', 'Trip Count', 'Total Fare', 'Total Tips']
            for col, header in enumerate(date_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.border = border
                cell.alignment = center_alignment
            
            # Group trips by date
            from collections import defaultdict
            date_summary = defaultdict(lambda: {'count': 0, 'fare': 0, 'tips': 0})
            
            for trip in data['trips']:
                trip_date = str(trip.start_date)
                date_summary[trip_date]['count'] += 1
                date_summary[trip_date]['fare'] += float(trip.trip_amount) if trip.trip_amount else 0
                date_summary[trip_date]['tips'] += float(trip.tips) if trip.tips else 0
            
            # Add date summary data
            for date_str, summary in sorted(date_summary.items()):
                row += 1
                date_data = [
                    date_str,
                    summary['count'],
                    summary['fare'],
                    summary['tips']
                ]
                
                for col, value in enumerate(date_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [3, 4] and isinstance(value, (int, float)):  # Fare and tips columns
                        cell.number_format = '$#,##0.00'
        
        # Notes Section
        row += 3
        ws[f'A{row}'] = "Notes"
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        
        row += 1
        ws[f'A{row}'] = "Driver signature is required upon collection. Please contact admin for any discrepancies."
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        row += 2
        ws[f'A{row}'] = f"Generated on: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        
        # Save and upload
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        key = f"dtr_receipts/driver_{data['driver'].id}/receipt_{data['receipt'].receipt_number}_{datetime.now(timezone.utc).timestamp()}.xlsx"
        s3_utils.upload_file(excel_bytes, key, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return key
    except Exception as e:
        logger.error("Error generating DTR Excel document: %s", e, exc_info=True)
        raise e