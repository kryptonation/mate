# app/utils/exporter_utils.py

import csv
import json
from io import BytesIO, StringIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from app.utils.logger import get_logger

logger = get_logger(__name__)

class BaseExporter:
    """Abstract base class for exporters."""
    def __init__(self, data: List[Dict[str, Any]]):
        if not data:
            raise ValueError("No data provided for export.")
        self.data = data
        self.headers = list(data[0].keys())

    def export(self) -> BytesIO:
        """Exports the data to a file-like object."""
        raise NotImplementedError

class ExcelExporter(BaseExporter):
    """Exports data to an Excel (XLSX) file in memory."""
    def export(self) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        
        # Style for header
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Write header
        for col_num, header_title in enumerate(self.headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            sheet.column_dimensions[cell.column_letter].width = 20

        # Write data rows
        for row_num, row_data in enumerate(self.data, 2):
            for col_num, header in enumerate(self.headers, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=str(row_data.get(header, "")))
                cell.border = thin_border
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

class CSVExporter(BaseExporter):
    """Exports data to a CSV file in memory."""
    def export(self) -> BytesIO:
        """Exports the data to a file-like object."""
        string_io = StringIO()
        writer = csv.DictWriter(string_io, fieldnames=self.headers)
        writer.writeheader()
        writer.writerows(self.data)
        
        # Convert StringIO to BytesIO for streaming response
        output = BytesIO(string_io.getvalue().encode('utf-8'))
        output.seek(0)
        return output

class PDFExporter(BaseExporter):
    """Exports data to a PDF file in memory."""
    def export(self) -> BytesIO:
        """Exports the data to a file-like object."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        
        # Prepare data for the table
        table_data = [self.headers]
        for row in self.data:
            table_data.append([str(row.get(header, "")) for header in self.headers])
            
        # Create table and apply style
        table = Table(table_data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)
        
        # Build the PDF
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Exported Data", styles['h1']))
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return buffer

class JSONExporter(BaseExporter):
    """Exports data to a JSON file in memory."""
    def export(self) -> BytesIO:
        """Exports the data to a file-like object."""
        json_string = json.dumps(self.data, indent=4, default=str) # Use default=str for dates
        output = BytesIO(json_string.encode('utf-8'))
        output.seek(0)
        return output

class ExporterFactory:
    """Factory to get the correct exporter based on the format."""
    
    @staticmethod
    def get_exporter(format_type: str, data: List[Dict[str, Any]]) -> BaseExporter:
        """
        Returns an instance of the appropriate exporter class based on the format type.
        """
        format_type = format_type.lower()
        if format_type == "excel":
            return ExcelExporter(data)
        if format_type == "csv":
            return CSVExporter(data)
        if format_type == "pdf":
            return PDFExporter(data)
        if format_type == "json":
            return JSONExporter(data)
        raise ValueError(f"Unsupported export format: {format_type}")