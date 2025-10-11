### app/periodic_reports/tasks.py

# Standard library imports
import os
import io
import json
import traceback
from datetime import datetime, date, timedelta
from typing import Dict, Any, Optional

# Third party imports
from celery import Celery
from sqlalchemy.orm import sessionmaker
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import pandas as pd

# Local imports
from app.core.config import settings
from app.core.db import engine
from app.utils.logger import get_logger
from app.worker.app import app as celery_app  # Use existing Celery app
from app.periodic_reports.models import (
    GeneratedReport, ReportConfiguration, ReportStatus, ReportType, ReportFormat
)
from app.periodic_reports.generators import (
    driver_summary_generator, medallion_status_generator,
    vehicle_inspection_generator, financial_summary_generator,
    lease_expiry_generator, violation_summary_generator,
    ezpass_transactions_generator, trip_analytics_generator,
    sla_performance_generator, audit_trail_summary_generator,
    payment_summary_generator
)
from app.utils.email_utils import send_email, send_simple_email
from app.utils.s3_utils import s3_utils  # Use existing S3 utilities

logger = get_logger(__name__)

# Create session
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@celery_app.task(bind=True, max_retries=3)
def generate_report_task(self, report_id: int, override_parameters: Dict[str, Any] = None, send_email: Optional[bool] = None):
    """
    Celery task to generate a report
    """
    db = SessionLocal()
    start_time = datetime.utcnow()
    
    try:
        # Get the generated report record
        generated_report = db.query(GeneratedReport).filter(
            GeneratedReport.id == report_id
        ).first()
        
        if not generated_report:
            logger.error("Generated report %d not found", report_id)
            return
        
        # Update status to processing
        generated_report.status = ReportStatus.PROCESSING
        db.commit()
        
        # Get the configuration
        config = generated_report.configuration
        if not config:
            raise ValueError(f"Report configuration not found for generated report {report_id}")
        
        logger.info("Starting report generation for %s (ID: %d)", config.name, report_id)

        # Merge parameters
        final_parameters = config.parameters.copy() if config.parameters else {}
        if override_parameters:
            final_parameters.update(override_parameters)
        
        # Add date range to parameters if provided
        if generated_report.report_period_start:
            final_parameters['start_date'] = generated_report.report_period_start.isoformat()
        if generated_report.report_period_end:
            final_parameters['end_date'] = generated_report.report_period_end.isoformat()
        
        # Generate the report data
        report_data = _generate_report_data(db, config.report_type, final_parameters)
        
        # Create the report file
        file_path = _create_report_file(
            config, generated_report, report_data, final_parameters
        )
        
        # Update the generated report record
        generated_report.status = ReportStatus.COMPLETED
        generated_report.file_path = file_path
        generated_report.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else None
        generated_report.execution_time_seconds = int((datetime.utcnow() - start_time).total_seconds())
        db.commit()
        
        # Send email if configured
        email_should_send = send_email if send_email is not None else config.auto_email
        if email_should_send and config.email_recipients:
            _send_report_email(db, generated_report, config)

        logger.info("Successfully generated report %s", generated_report.report_name)

    except Exception as e:
        logger.error("Error generating report %d: %s", report_id, str(e), exc_info=True)

        # Update status to failed
        if 'generated_report' in locals():
            generated_report.status = ReportStatus.FAILED
            generated_report.error_message = str(e)
            generated_report.execution_time_seconds = int((datetime.utcnow() - start_time).total_seconds())
            db.commit()
        
        # Retry if not at max retries
        if self.request.retries < self.max_retries:
            logger.info("Retrying report generation %d (attempt %d)", report_id, self.request.retries + 1)
            raise self.retry(countdown=60 * (self.request.retries + 1))
        
    finally:
        db.close()


@celery_app.task
def generate_scheduled_reports_task():
    """
    Celery task to check for and generate scheduled reports
    """
    db = SessionLocal()
    
    try:
        from app.periodic_reports.services import periodic_reports_service
        
        # Get reports due for generation
        due_reports = periodic_reports_service.get_scheduled_reports_due(db)

        logger.info("Found %d reports due for generation", len(due_reports))

        for config in due_reports:
            try:
                # Create a generated report record
                generated_report = GeneratedReport(
                    configuration_id=config.id,
                    report_name=periodic_reports_service._generate_report_name(config, None, None),
                    generation_date=datetime.utcnow(),
                    status=ReportStatus.PENDING,
                    created_by=config.created_by
                )
                
                db.add(generated_report)
                db.commit()
                db.refresh(generated_report)
                
                # Queue the report generation
                generate_report_task.delay(generated_report.id)
                
                logger.info(f"Queued scheduled report: {config.name}")
                
            except Exception as e:
                logger.error(f"Error queuing scheduled report {config.name}: {str(e)}")
                continue
                
    except Exception as e:
        logger.error(f"Error in scheduled reports task: {str(e)}", exc_info=True)
    finally:
        db.close()


def _generate_report_data(db, report_type: ReportType, parameters: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate report data based on report type and parameters
    """
    generators = {
        ReportType.DRIVER_SUMMARY: driver_summary_generator,
        ReportType.MEDALLION_STATUS: medallion_status_generator,
        ReportType.VEHICLE_INSPECTION: vehicle_inspection_generator,
        ReportType.FINANCIAL_SUMMARY: financial_summary_generator,
        ReportType.LEASE_EXPIRY: lease_expiry_generator,
        ReportType.VIOLATION_SUMMARY: violation_summary_generator,
        ReportType.EZPASS_TRANSACTIONS: ezpass_transactions_generator,
        ReportType.TRIP_ANALYTICS: trip_analytics_generator,
        ReportType.SLA_PERFORMANCE: sla_performance_generator,
        ReportType.AUDIT_TRAIL_SUMMARY: audit_trail_summary_generator,
        ReportType.PAYMENT_SUMMARY: payment_summary_generator
    }
    
    generator_func = generators.get(report_type)
    if not generator_func:
        raise ValueError(f"No generator found for report type: {report_type}")
    
    return generator_func.generate(db, parameters)


def _create_report_file(
    config: ReportConfiguration,
    generated_report: GeneratedReport,
    report_data: Dict[str, Any],
    parameters: Dict[str, Any]
) -> str:
    """
    Create the actual report file based on the format and upload to S3
    """
    # Ensure output directory exists (for local backup/temp storage)
    output_dir = config.output_directory or os.path.join(settings.document_storage_dir or "/tmp", "reports")
    os.makedirs(output_dir, exist_ok=True)
    
    # Create filename
    filename = f"{generated_report.report_name}.{config.output_format.value}"
    local_file_path = os.path.join(output_dir, filename)
    
    # Generate the report file locally first
    if config.output_format == ReportFormat.PDF:
        _create_pdf_report(local_file_path, config, report_data, parameters)
    elif config.output_format == ReportFormat.EXCEL:
        _create_excel_report(local_file_path, config, report_data, parameters)
    elif config.output_format == ReportFormat.CSV:
        _create_csv_report(local_file_path, config, report_data, parameters)
    elif config.output_format == ReportFormat.JSON:
        _create_json_report(local_file_path, config, report_data, parameters)
    else:
        raise ValueError(f"Unsupported output format: {config.output_format}")
    
    # Upload to S3
    s3_key = f"reports/{config.report_type.value}/{generated_report.generation_date.strftime('%Y/%m/%d')}/{filename}"
    
    try:
        with open(local_file_path, 'rb') as file_obj:
            upload_success = s3_utils.upload_file(
                file_obj=file_obj,
                key=s3_key,
                content_type=_get_content_type(config.output_format)
            )
        
        if upload_success:
            logger.info(f"Successfully uploaded report to S3: {s3_key}")
            
            # Generate presigned URL for download (valid for 7 days)
            presigned_url = s3_utils.generate_presigned_url(s3_key, expiration=7*24*3600)
            
            # Clean up local file after successful upload
            try:
                os.remove(local_file_path)
                logger.info(f"Cleaned up local file: {local_file_path}")
            except Exception as e:
                logger.warning(f"Could not remove local file {local_file_path}: {str(e)}")
            
            # Return S3 key as the file path for database storage
            return s3_key
        else:
            logger.error(f"Failed to upload report to S3: {s3_key}")
            # Fall back to local file path if S3 upload fails
            return local_file_path
            
    except Exception as e:
        logger.error(f"Error uploading report to S3: {str(e)}")
        # Fall back to local file path if S3 upload fails
        return local_file_path


def _get_content_type(output_format: ReportFormat) -> str:
    """Get the appropriate content type for the output format"""
    content_types = {
        ReportFormat.PDF: "application/pdf",
        ReportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ReportFormat.CSV: "text/csv",
        ReportFormat.JSON: "application/json"
    }
    return content_types.get(output_format, "application/octet-stream")


def _create_pdf_report(file_path: str, config: ReportConfiguration, report_data: Dict[str, Any], parameters: Dict[str, Any]):
    """Create a PDF report"""
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Center
    )
    story.append(Paragraph(config.name, title_style))
    story.append(Spacer(1, 20))
    
    # Report info
    info_data = [
        ['Report Type:', config.report_type.value.replace('_', ' ').title()],
        ['Generated On:', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')],
        ['Period:', f"{parameters.get('start_date', 'N/A')} to {parameters.get('end_date', 'N/A')}"]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Add data tables
    if 'tables' in report_data:
        for table_name, table_data in report_data['tables'].items():
            # Table title
            story.append(Paragraph(table_name.replace('_', ' ').title(), styles['Heading2']))
            story.append(Spacer(1, 10))
            
            if table_data and len(table_data) > 0:
                # Convert to table format
                headers = list(table_data[0].keys())
                data = [headers]
                for row in table_data:
                    data.append([str(row.get(header, '')) for header in headers])
                
                # Create table
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(table)
            else:
                story.append(Paragraph("No data available", styles['Normal']))
            
            story.append(Spacer(1, 20))
    
    # Add summary if available
    if 'summary' in report_data:
        story.append(Paragraph("Summary", styles['Heading2']))
        for key, value in report_data['summary'].items():
            story.append(Paragraph(f"{key}: {value}", styles['Normal']))
    
    doc.build(story)


def _create_excel_report(file_path: str, config: ReportConfiguration, report_data: Dict[str, Any], parameters: Dict[str, Any]):
    """Create a comprehensive Excel report with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, Reference
    import pandas as pd
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 1. Summary Sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Report Information"])
    summary_ws.append(["Report Name", config.name])
    summary_ws.append(["Report Type", config.report_type.value.replace('_', ' ').title()])
    summary_ws.append(["Generated On", datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')])
    summary_ws.append(["Period Start", str(parameters.get('start_date', 'N/A'))])
    summary_ws.append(["Period End", str(parameters.get('end_date', 'N/A'))])
    summary_ws.append([])  # Empty row
    
    # Add summary statistics if available
    if 'summary' in report_data:
        summary_ws.append(["Summary Statistics"])
        for key, value in report_data['summary'].items():
            if key not in ['period_start', 'period_end']:  # Skip duplicate date fields
                display_key = key.replace('_', ' ').title()
                summary_ws.append([display_key, str(value)])
    
    # Format summary sheet
    for row in summary_ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Auto-adjust column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Data Sheets
    if 'tables' in report_data:
        for table_name, table_data in report_data['tables'].items():
            if table_data:
                # Create sheet with proper name
                sheet_name = table_name.replace('_', ' ').title()[:31]  # Excel sheet name limit
                ws = wb.create_sheet(sheet_name)
                
                # Convert to DataFrame for better handling
                df = pd.DataFrame(table_data)
                
                # Add title row
                ws.append([sheet_name])
                ws.append([])  # Empty row
                
                # Add column headers
                headers = list(df.columns)
                ws.append(headers)
                
                # Add data rows
                for row in df.values:
                    # Convert any datetime objects to strings
                    formatted_row = []
                    for cell_value in row:
                        if isinstance(cell_value, (datetime, date)):
                            formatted_row.append(cell_value.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            formatted_row.append(str(cell_value) if cell_value is not None else '')
                    ws.append(formatted_row)
                
                # Format headers
                header_row = 3  # Row with column headers
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                
                # Format title
                title_cell = ws.cell(row=1, column=1)
                title_cell.font = Font(bold=True, size=16)
                title_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                # Add borders to data
                max_row = ws.max_row
                max_col = len(headers)
                for row in ws.iter_rows(min_row=header_row, max_row=max_row, 
                                      min_col=1, max_col=max_col):
                    for cell in row:
                        cell.border = border
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Add simple chart if numerical data exists
                try:
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0 and len(df) <= 50:  # Only for manageable datasets
                        chart = BarChart()
                        chart.title = f"{sheet_name} Summary"
                        chart.y_axis.title = 'Values'
                        chart.x_axis.title = 'Categories'
                        
                        # Add data to chart (first numeric column)
                        data = Reference(ws, min_col=headers.index(numeric_cols[0]) + 1, 
                                       min_row=header_row, max_row=min(max_row, header_row + 20))
                        categories = Reference(ws, min_col=1, min_row=header_row + 1, 
                                             max_row=min(max_row, header_row + 20))
                        
                        chart.add_data(data, titles_from_data=False)
                        chart.set_categories(categories)
                        chart.height = 10
                        chart.width = 15
                        
                        # Position chart to the right of data
                        ws.add_chart(chart, f"{chr(65 + max_col + 1)}{header_row}")
                except Exception as e:
                    # Chart creation is optional, continue without it
                    logger.warning(f"Could not create chart for {sheet_name}: {str(e)}")
    
    # 3. Charts and Analytics Sheet (if summary data exists)
    if 'summary' in report_data and len(report_data['summary']) > 2:
        try:
            analytics_ws = wb.create_sheet("Analytics")
            analytics_ws.append(["Report Analytics Dashboard"])
            analytics_ws.append([])
            
            # Create a summary table for chart data
            chart_data = []
            for key, value in report_data['summary'].items():
                if isinstance(value, (int, float)) and key not in ['period_start', 'period_end']:
                    chart_data.append([key.replace('_', ' ').title(), value])
            
            if chart_data:
                analytics_ws.append(["Metric", "Value"])
                for row in chart_data:
                    analytics_ws.append(row)
                
                # Create a simple bar chart
                chart = BarChart()
                chart.title = "Report Summary Metrics"
                chart.y_axis.title = 'Values'
                chart.x_axis.title = 'Metrics'
                
                data = Reference(analytics_ws, min_col=2, min_row=3, max_row=3 + len(chart_data))
                categories = Reference(analytics_ws, min_col=1, min_row=4, max_row=3 + len(chart_data))
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(categories)
                chart.height = 15
                chart.width = 20
                
                analytics_ws.add_chart(chart, "D3")
                
        except Exception as e:
            logger.warning(f"Could not create analytics sheet: {str(e)}")
    
    # Save the workbook
    wb.save(file_path)
    logger.info(f"Enhanced Excel report created: {file_path}")


def _create_csv_report(file_path: str, config: ReportConfiguration, report_data: Dict[str, Any], parameters: Dict[str, Any]):
    """Create a CSV report"""
    # For CSV, combine all tables into one file
    all_data = []
    
    # Add header information
    all_data.append({
        'Report Name': config.name,
        'Report Type': config.report_type.value,
        'Generated On': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC'),
        'Period Start': parameters.get('start_date', 'N/A'),
        'Period End': parameters.get('end_date', 'N/A')
    })
    
    # Add empty row
    all_data.append({})
    
    # Add data from tables
    if 'tables' in report_data:
        for table_name, table_data in report_data['tables'].items():
            # Add table header
            all_data.append({'Table': table_name})
            if table_data:
                all_data.extend(table_data)
            all_data.append({})  # Empty row between tables
    
    df = pd.DataFrame(all_data)
    df.to_csv(file_path, index=False)


def _create_json_report(file_path: str, config: ReportConfiguration, report_data: Dict[str, Any], parameters: Dict[str, Any]):
    """Create a JSON report"""
    output_data = {
        'report_info': {
            'name': config.name,
            'type': config.report_type.value,
            'generated_on': datetime.utcnow().isoformat(),
            'parameters': parameters
        },
        'data': report_data
    }
    
    with open(file_path, 'w') as f:
        json.dump(output_data, f, indent=2, default=str)


def _send_report_email(db, generated_report: GeneratedReport, config: ReportConfiguration):
    """Send the generated report via email"""
    try:
        # Check if file is stored in S3 (S3 keys start with "reports/")
        is_s3_file = generated_report.file_path and generated_report.file_path.startswith("reports/")
        
        if is_s3_file:
            # Generate presigned URL for S3 files (valid for 7 days)
            download_url = s3_utils.generate_presigned_url(
                generated_report.file_path, 
                expiration=7*24*3600  # 7 days
            )
            
            subject = f"Report: {config.name}"
            body = f"""
Your scheduled report "{config.name}" has been generated successfully.

Report Details:
- Type: {config.report_type.value.replace('_', ' ').title()}
- Generated On: {generated_report.generation_date.strftime('%Y-%m-%d %H:%M:%S UTC')}
- Period: {generated_report.report_period_start or 'N/A'} to {generated_report.report_period_end or 'N/A'}
- File Size: {generated_report.file_size or 0} bytes

Download Link (valid for 7 days): {download_url or 'Download link could not be generated'}

Note: This download link will expire in 7 days for security purposes.
            """.strip()
            
            # Send email without attachment (using download link instead)
            from app.utils.email_utils import send_simple_email
            success = send_simple_email(
                recipients=config.email_recipients,
                subject=subject,
                body=body
            )
            
        else:
            # Handle local files (fallback)
            subject = f"Report: {config.name}"
            body = f"""
Your scheduled report "{config.name}" has been generated successfully.

Report Details:
- Type: {config.report_type.value.replace('_', ' ').title()}
- Generated On: {generated_report.generation_date.strftime('%Y-%m-%d %H:%M:%S UTC')}
- Period: {generated_report.report_period_start or 'N/A'} to {generated_report.report_period_end or 'N/A'}
- File Size: {generated_report.file_size or 0} bytes

Please find the report attached.
            """.strip()
            
            # Send email with attachment for local files
            success = send_email(
                recipients=config.email_recipients,
                subject=subject,
                body=body,
                attachments=[generated_report.file_path] if generated_report.file_path and os.path.exists(generated_report.file_path) else None
            )
        
        if success:
            # Update email status
            generated_report.email_sent = True
            generated_report.email_sent_at = datetime.utcnow()
            db.commit()
            logger.info(f"Email sent for report {generated_report.report_name}")
        else:
            generated_report.email_error = "Failed to send email"
            db.commit()
            logger.error(f"Failed to send email for report {generated_report.report_name}")
        
    except Exception as e:
        logger.error(f"Error sending email for report {generated_report.report_name}: {str(e)}")
        generated_report.email_error = str(e)
        db.commit()


@celery_app.task
def cleanup_old_reports_task():
    """
    Celery task to cleanup old report files (both local and S3)
    """
    db = SessionLocal()
    
    try:
        # Delete report files older than 90 days
        cutoff_date = datetime.utcnow() - timedelta(days=90)
        
        old_reports = db.query(GeneratedReport).filter(
            GeneratedReport.generation_date < cutoff_date,
            GeneratedReport.status == ReportStatus.COMPLETED
        ).all()
        
        deleted_count = 0
        s3_deleted_count = 0
        
        for report in old_reports:
            if report.file_path:
                # Check if it's an S3 key (starts with "reports/")
                if report.file_path.startswith("reports/"):
                    # This is an S3 key, delete from S3
                    try:
                        if s3_utils.delete_file(report.file_path):
                            logger.info(f"Deleted S3 file: {report.file_path}")
                            s3_deleted_count += 1
                        else:
                            logger.warning(f"Failed to delete S3 file: {report.file_path}")
                        
                        # Clear the file path regardless of deletion success
                        report.file_path = None
                        
                    except Exception as e:
                        logger.error(f"Error deleting S3 file {report.file_path}: {str(e)}")
                        
                # Check if it's a local file path
                elif os.path.exists(report.file_path):
                    try:
                        os.remove(report.file_path)
                        report.file_path = None
                        deleted_count += 1
                        logger.info(f"Deleted local file: {report.file_path}")
                    except Exception as e:
                        logger.error(f"Error deleting local file {report.file_path}: {str(e)}")
                else:
                    # File path exists but file doesn't exist (cleanup stale reference)
                    report.file_path = None
        
        db.commit()
        logger.info(f"Cleaned up {deleted_count} local files and {s3_deleted_count} S3 files")
        
    except Exception as e:
        logger.error(f"Error in cleanup task: {str(e)}")
    finally:
        db.close()


@celery_app.task
def send_weekly_summary_task():
    """
    Send weekly summary of report generation activities
    """
    db = SessionLocal()
    
    try:
        # Get reports generated in the last week
        week_start = datetime.utcnow() - timedelta(days=7)
        
        reports_this_week = db.query(GeneratedReport).filter(
            GeneratedReport.generation_date >= week_start
        ).all()
        
        # Calculate statistics
        total_reports = len(reports_this_week)
        successful_reports = len([r for r in reports_this_week if r.status == ReportStatus.COMPLETED])
        failed_reports = len([r for r in reports_this_week if r.status == ReportStatus.FAILED])
        
        # Group by report type
        type_breakdown = {}
        for report in reports_this_week:
            if report.configuration and report.configuration.report_type:
                report_type = report.configuration.report_type.value
                if report_type not in type_breakdown:
                    type_breakdown[report_type] = {'total': 0, 'successful': 0, 'failed': 0}
                type_breakdown[report_type]['total'] += 1
                if report.status == ReportStatus.COMPLETED:
                    type_breakdown[report_type]['successful'] += 1
                elif report.status == ReportStatus.FAILED:
                    type_breakdown[report_type]['failed'] += 1
        
        # Create summary email
        summary_text = f"""
        Weekly Report Generation Summary
        ===============================
        
        Period: {week_start.strftime('%Y-%m-%d')} to {datetime.utcnow().strftime('%Y-%m-%d')}
        
        Overall Statistics:
        - Total Reports Generated: {total_reports}
        - Successful: {successful_reports}
        - Failed: {failed_reports}
        - Success Rate: {(successful_reports/total_reports*100) if total_reports > 0 else 0:.1f}%
        
        Breakdown by Report Type:
        """
        
        for report_type, stats in type_breakdown.items():
            summary_text += f"""
        {report_type.replace('_', ' ').title()}:
          - Total: {stats['total']}
          - Successful: {stats['successful']}
          - Failed: {stats['failed']}
            """
        
        # Send to configured administrators (you might want to create an admin email list)
        admin_emails = ["admin@bigappletaxi.com"]  # Configure this in settings
        
        if admin_emails:
            send_simple_email(
                recipients=admin_emails,
                subject="BAT Weekly Report Generation Summary",
                body=summary_text
            )
        
        logger.info(f"Sent weekly summary: {total_reports} reports generated this week")
        
    except Exception as e:
        logger.error(f"Error sending weekly summary: {str(e)}")
    finally:
        db.close()


@celery_app.task
def retry_failed_reports_task():
    """
    Retry failed report generations that haven't been retried recently
    """
    db = SessionLocal()
    
    try:
        # Find failed reports from the last 24 hours that haven't been retried recently
        cutoff_time = datetime.utcnow() - timedelta(hours=24)
        retry_cutoff = datetime.utcnow() - timedelta(hours=2)  # Don't retry more than once every 2 hours
        
        failed_reports = db.query(GeneratedReport).filter(
            GeneratedReport.status == ReportStatus.FAILED,
            GeneratedReport.generation_date >= cutoff_time,
            GeneratedReport.updated_on <= retry_cutoff  # Haven't been updated recently
        ).limit(10).all()  # Limit to 10 retries at a time
        
        retried_count = 0
        for report in failed_reports:
            try:
                # Reset status and retry
                report.status = ReportStatus.PENDING
                report.error_message = None
                db.commit()
                
                # Queue the report generation again
                generate_report_task.delay(report.id)
                retried_count += 1
                
                logger.info(f"Retrying failed report: {report.report_name}")
                
            except Exception as e:
                logger.error(f"Error retrying report {report.id}: {str(e)}")
                continue
        
        if retried_count > 0:
            logger.info(f"Retried {retried_count} failed reports")
        
    except Exception as e:
        logger.error(f"Error in retry failed reports task: {str(e)}")
    finally:
        db.close()
