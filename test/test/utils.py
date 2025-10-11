import os
import openpyxl
from datetime import datetime
import mimetypes
import logging
import json
from test.test_db import test_step
from sqlalchemy.orm import Session
from fastapi.testclient import TestClient
from app.bpm.utils import fetch_case_entity
from app.bat.utils import fetch_driver_by_id

logger = logging.getLogger(__name__)

class TestingFlows:
    def __init__(self, client, db_session):
        self.client = client
        self.db_session = db_session
        self.case_no = ""
        self.session_id = None

    def login(self, login_data):
        logger.info("Attempting login")
        response = self.client.post("/login", json=login_data.dict())
        assert response.status_code == 200, f"Login failed: {response.status_code}"
        self.session_id = response.cookies.get("bat_session_id")
        assert self.session_id, "Session ID not set in cookies"
        logger.info("Login successful. Session ID: %s", self.session_id)

    def create_case(self, case_data):
        logger.info("Creating case")
        if not self.session_id:
            logger.error("Session ID is not set. Please login first.")
            raise ValueError("Session ID is not set. Please login first.")
        
        self.client.cookies.set("bat_session_id", self.session_id)
        response = self.client.post("/case", json=case_data)
        assert response.status_code == 200, f"Case creation failed: {response.status_code}, {response.json()}"
        self.case_no = response.json().get("case_no")
        assert self.case_no, "Case number not returned"
        logger.info("Case created successfully. Case No: %s", self.case_no)

    def fetch_step_data(self, step_id: str, case_params: dict = None):
        logger.info("Fetching step data for step %s", step_id)
        url = f"/case/{self.case_no}/{step_id}"
        response = self.client.get(url, params=case_params) if case_params else self.client.get(url)
        #assert response.status_code == 200, f"Fetch for step {step_id} failed: {response.json().get('detail')}"
        return response

    def process_step(self, step_id: str, payload):
        logger.info("Processing step %s", step_id)
        response = self.client.post(f"/case/{self.case_no}", json=payload)
        return response

    def move_step(self):
        logger.info("Moving to the next step for case %s", self.case_no)
        response = self.client.post(f"/case/{self.case_no}/move")
        assert response.status_code == 200, "Failed to move step"
        logger.info("Step moved successfully.")

    def upload_document(self, step_id, payload, document_path):
        logger.info("Uploading document for step %s", step_id)
        document_path = os.path.abspath(document_path)
        assert os.path.exists(document_path), f"Document not found: {document_path}"
        mime_type, _ = mimetypes.guess_type(document_path)
        if not mime_type:
            raise ValueError(f"Could not determine MIME type for file: {document_path}")
        with open(document_path, "rb") as file:
            files = {"file": (os.path.basename(document_path), file, mime_type)}
            response = self.client.post("/upload-document", data=payload, files=files)
        return response



class TestingBase:
    def __init__(self, client, db_session):
        self.client = client
        self.db_session = db_session
        self.session_id = None

    def login(self, login_data):
        logger.info("Attempting login")
        response = self.client.post("/login", json=login_data.dict())
        assert response.status_code == 200, f"Login failed: {response.status_code}"
        self.session_id = response.cookies.get("bat_session_id")
        assert self.session_id, "Session ID not set in cookies"
        logger.info("Login successful. Session ID: %s", self.session_id)

class TestingMedallionOwnerListing(TestingBase):
    def search_medallion_owner(self, search_data):
        logger.info(f"Fetching medallion owner with search data: {search_data}")
        if search_data != {}:
            response = self.client.get("/api/owner-listing/v2", params=search_data)
        else:
            response = self.client.get("/api/owner-listing/v2")
        response_data = response.json()
        logger.info("Response data of get medallion owner is :\n%s", json.dumps(response_data))
        return response

class TestingSearchDriver(TestingBase):
    def search_driver(self, search_data):
        logger.info(f"Fetching driver with search data: {search_data}")
        if search_data != {}:
            response = self.client.get("/drivers", params=search_data)
        else:
            response = self.client.get("/drivers")
        response_data = response.json()
        logger.info("Response data of get driver is :\n%s", json.dumps(response_data))
        return response
    

class TestingVehicleEntitySearch(TestingBase):
    def search_vehicle_entity(self, search_data):
        logger.info(f"Fetching vehicle entity with search data: {search_data}")
        if search_data != {}:
            response = self.client.get("/entities", params=search_data)
        else:
            response = self.client.get("/entities")
        response_data = response.json()
        logger.info("Response data of get vehicle entity is :\n%s", json.dumps(response_data))
        return response
    
class TestingManageVehicleList(TestingBase):
    def search_manage_vehicle_list(self, search_data):
        logger.info(f"fatching manage vehicle list data:{search_data}")

        if search_data !={} :
            response= self.client.get("/vehicles" , params=search_data)
        else :
            response= self.client.get("/vehicles")
        response_data= response.json()
        logger.info("response data of get manage vehicele list :\n%s",json.dumps(response_data))
        return response


class TestingListMedallions(TestingBase):
    def search_medallion(self, search_data):
        logger.info(f"Fetching medallions with: {search_data}")
        if search_data != {}:
            response = self.client.get("/medallions", params=search_data)
        else:
            response = self.client.get("/medallions")
        response_data = response.json()
        logger.info("Response data of get medallions is :\n%s", json.dumps(response_data))
        return response


class TestingDocumentUpload(TestingBase):

    def upload_document(self, payload):
        logger.info("Uploading document")
        # Extract document_path from the payload and remove it from the data
        document_path = payload.pop('document_path', None)
        if not document_path:
            raise ValueError("Document path is required")
        document_path = os.path.abspath(document_path)
        assert os.path.exists(document_path), f"Document not found: {document_path}"
        mime_type, _ = mimetypes.guess_type(document_path)
        if not mime_type:
            raise ValueError(f"Could not determine MIME type for file: {document_path}")
        with open(document_path, "rb") as file:
            files = {"file": (os.path.basename(document_path), file, mime_type)}
            response = self.client.post("/upload-document", data=payload, files=files)
        return response
    

class TestingDocumentDelete(TestingBase):

    def delete_document(self, payload):
        logger.info("Deleting document")

        # Extract document_id from payload
        document_id = payload.get("document_id")
        if not document_id:
            raise ValueError("document_id is required in the payload")
        response = self.client.delete(f"/delete-document/{document_id}")
        return response
    

class TestingDeactivateMedallions(TestingBase):

    def deactivate_medallions(self, payload):
        response = self.client.put("/medallions/deactivate", json=payload)
        return response
    

def deep_compare(expected, actual, key_prefix="", matched_keys=None):
    if matched_keys is None:
        matched_keys = []

    if isinstance(expected, dict) and isinstance(actual, dict):
        for key, expected_value in expected.items():
            new_prefix = f"{key_prefix}.{key}" if key_prefix else key

            if key not in actual:
                logger.error(f"❌ Missing key: {new_prefix}")
                return f"Key {new_prefix} is missing in actual response"

            result = deep_compare(expected_value, actual[key], new_prefix, matched_keys)
            if result:
                return result

    elif isinstance(expected, list) and isinstance(actual, list):
        if len(expected) != len(actual):
            return f"❌ List length mismatch at {key_prefix}: Expected {len(expected)} items, got {len(actual)} items"

        # Sort lists to ignore order
        expected_sorted = sorted(expected, key=lambda x: str(x) if isinstance(x, dict) else str(x))
        actual_sorted = sorted(actual, key=lambda x: str(x) if isinstance(x, dict) else str(x))

        for i, (exp_item, act_item) in enumerate(zip(expected_sorted, actual_sorted)):
            result = deep_compare(exp_item, act_item, f"{key_prefix}[{i}]", matched_keys)
            if result:
                return result

    else:
        if expected == actual:
            logger.info(f"✅ Matching key: {key_prefix} | Value: {expected}")
            matched_keys.append(key_prefix)
        else:
            logger.error(f"❌ Mismatch at {key_prefix}: Expected={expected}, Actual={actual}")
            return f"Mismatch at {key_prefix}: Expected={expected}, Actual={actual}"

    return None


def compare_response(expected_response, actual_response):
    matched_keys = []
    result = deep_compare(expected_response, actual_response, matched_keys=matched_keys)

    if result:
        logger.error(f"Test failed: {result}")
        return f"Fail - {result}"
    else:
        logger.info("✅ All expected keys matched successfully")
        logger.info(f"Matched Keys: {matched_keys}")
        return "Pass"

def update_excel_outcome(file_path, test_case_id, outcome_text):
    """
    Update the 'Outcome' field in the Excel file for a given TestCaseID.
    - Headers are in row 1.
    - Test case data starts from row 2.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Get the current timestamp
    current_timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    # Locate the "Outcome" column (headers are in row 1)
    header_row = [cell.value for cell in sheet[4]]  # Read headers from row 1
    try:
        outcome_col_idx = header_row.index("Outcome") + 1  # Convert to 1-based index
    except ValueError:
        raise ValueError("Column 'Outcome' not found in the Excel sheet.")

    # Find the correct row for the TestCaseID
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # Start from row 2
        if row[0].value == test_case_id:  # Assuming "TestCaseID" is in column 1
            row[outcome_col_idx - 1].value = f"{outcome_text} ({current_timestamp})"  # Update "Outcome"
            break

    wb.save(file_path)
    wb.close()

def update_excel_outcome_flows(file_path, sheet_name, test_case_id, outcome_text):
    """
    Update the 'Outcome' field in the Excel file for a given TestCaseID.
    - Headers are in row 4.
    - Test case data starts from row 5.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # Get the current timestamp
    current_timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    # Locate the "Outcome" column (headers are in row 4)
    header_row = [cell.value for cell in sheet[4]]  # Read headers from row 4
    try:
        outcome_col_idx = header_row.index("Outcome") + 1  # Convert to 1-based index
    except ValueError:
        raise ValueError("Column 'Outcome' not found in the Excel sheet.")

    # Find the correct row for the TestCaseID
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):  # Start from row 5
        if row[0].value == test_case_id:  # Assuming "TestCaseID" is in column 1
            row[outcome_col_idx - 1].value = F"{outcome_text} ({current_timestamp})"  # Update "Outcome"
            break

    wb.save(file_path)
    wb.close()
